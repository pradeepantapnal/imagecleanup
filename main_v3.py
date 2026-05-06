"""
Intelligent Photo Cleanup Engine v3.2
======================================
Single unified solution — one script, one command, one output.

  python main_v3.py --folder "D:\\Photos"
  python main_v3.py --folder "D:\\Photos" --enable-vision --vision-limit 200
  python main_v3.py --folder "D:\\Photos" --dry-run
  python main_v3.py --folder "D:\\Photos" --limit 50 --enable-vision

All metrics computed inline during Stage 1 (blur, BRISQUE-approx, composite,
colorfulness, brightness, contrast, resolution). No intermediate files.

Output: 6-column Excel (Filename, Score/100, Decision, Category, Caption, Reason)
sorted worst-first, color-coded. Also CSV.

Optional:
  --enable-vision     Vision LLM on ambiguous photos (score 35-65 band)
  --enable-faces      Face detection/clustering
  --limit N           Process N images only
  --dry-run           Score distribution preview, no files written
  --vision-limit N    Max images to vision LLM (default: 20)
  --vision-model M    Primary Ollama model (default: llava)
  --metrics-excel X   Use external metrics workbook
  --generate-metrics  Generate metrics workbook and exit
  --output-dir DIR    Output directory (default: output)
  --clear-cache       Wipe cache before running
"""

import os
import re
import sys
import csv
import json
import math
import base64
import hashlib
import importlib.util
import sqlite3
import argparse
import logging
import requests
import asyncio
import platform
import shutil
import subprocess
from concurrent.futures import ProcessPoolExecutor, as_completed
from time import perf_counter, process_time
import gc
from io import BytesIO
from pathlib import Path
from datetime import datetime
from contextlib import contextmanager
from collections import defaultdict, Counter
from dataclasses import dataclass, asdict
from typing import Dict, List, Optional, Tuple

# ── Required ───────────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

try:
    from PIL import Image
except ImportError:
    print("pip install pillow"); sys.exit(1)

try:
    import aiohttp
except ImportError:
    print("pip install aiohttp"); sys.exit(1)

try:
    import imagehash
except ImportError:
    print("pip install imagehash"); sys.exit(1)

# ── Optional ───────────────────────────────────────────────────────────────────
try:
    import cv2
    HAS_CV2 = True
except ImportError:
    HAS_CV2 = False

try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False
    def tqdm(it, **kw):
        return it

HAS_CLIP = False
HAS_NVML = False
try:
    from pynvml import (
        nvmlInit, nvmlShutdown, nvmlDeviceGetHandleByIndex,
        nvmlDeviceGetUtilizationRates, nvmlDeviceGetMemoryInfo,
        nvmlDeviceGetPcieThroughput, NVML_PCIE_UTIL_TX_BYTES, NVML_PCIE_UTIL_RX_BYTES,
        nvmlDeviceGetPowerUsage, nvmlDeviceGetTemperature, NVML_TEMPERATURE_GPU,
    )
    HAS_NVML = True
except Exception:
    HAS_NVML = False

HAS_FACES = False

OLLAMA_NUM_PARALLEL = 4

# ── Constants ──────────────────────────────────────────────────────────────────
VERSION          = "3.2"
SUPPORTED_EXTS   = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tiff", ".tif", ".heic"}
DUPLICATE_THRESH = 8
BLUR_THRESH      = 80.0
EVENT_GAP_HOURS  = 6
OLLAMA_URL       = "http://localhost:11434/api/generate"
VISION_TIMEOUT   = 300
VISION_BAND_LOW  = 45
VISION_BAND_HIGH = 55


def _read_config(path: str) -> dict:
    cfg_path = Path(path)
    if not cfg_path.exists():
        return {}
    with cfg_path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, dict):
        raise ValueError(f"Config must be a JSON object: {path}")
    return data


def _cfg_get(cfg: dict, *keys, default=None):
    cur = cfg
    for k in keys:
        if not isinstance(cur, dict) or k not in cur:
            return default
        cur = cur[k]
    return cur


def _apply_config_defaults(args: argparse.Namespace, cfg: dict) -> argparse.Namespace:
    # Core paths/options
    if args.metrics_excel is None:
        args.metrics_excel = _cfg_get(cfg, "metrics_excel")
    if args.output_dir is None:
        args.output_dir = _cfg_get(cfg, "output_dir", default="output")
    if args.limit is None:
        args.limit = _cfg_get(cfg, "limit")

    # Feature toggles + vision options (supports legacy nested schema)
    if args.enable_vision is None:
        args.enable_vision = bool(_cfg_get(cfg, "enable_vision", default=_cfg_get(cfg, "v2_features", "enable_vision", default=False)))
    if args.enable_faces is None:
        args.enable_faces = bool(_cfg_get(cfg, "enable_faces", default=_cfg_get(cfg, "v2_features", "enable_faces", default=False)))
    if args.vision_model is None:
        args.vision_model = _cfg_get(cfg, "vision_model", default=_cfg_get(cfg, "v2_features", "vision_model", default="llava"))
    if args.vision_fallback is None:
        args.vision_fallback = _cfg_get(cfg, "vision_fallback", default=_cfg_get(cfg, "v2_features", "vision_fallback", default="llama3.2-vision"))
    if args.vision_limit is None:
        args.vision_limit = _cfg_get(cfg, "vision_limit", default=_cfg_get(cfg, "v2_features", "vision_max_images", default=20))

    # Execution toggles
    if args.dry_run is None:
        args.dry_run = bool(_cfg_get(cfg, "dry_run", default=False))
    if args.clear_cache is None:
        args.clear_cache = bool(_cfg_get(cfg, "clear_cache", default=False))

    # Global thresholds / model runtime settings
    global DUPLICATE_THRESH, EVENT_GAP_HOURS, BLUR_THRESH, OLLAMA_URL, VISION_TIMEOUT
    DUPLICATE_THRESH = int(_cfg_get(cfg, "duplicate_threshold", default=DUPLICATE_THRESH))
    EVENT_GAP_HOURS = int(_cfg_get(cfg, "event_gap_hours", default=EVENT_GAP_HOURS))
    BLUR_THRESH = float(_cfg_get(cfg, "blur_threshold", default=_cfg_get(cfg, "processing", "blur_threshold", default=BLUR_THRESH)))
    OLLAMA_URL = str(_cfg_get(cfg, "ollama_url", default=_cfg_get(cfg, "models", "ollama_url", default=OLLAMA_URL)))
    VISION_TIMEOUT = int(_cfg_get(cfg, "vision_timeout", default=_cfg_get(cfg, "models", "vision_timeout", default=VISION_TIMEOUT)))

    if args.output_dir is None:
        args.output_dir = "output"
    if args.vision_model is None:
        args.vision_model = "llava"
    if args.vision_fallback is None:
        args.vision_fallback = "llama3.2-vision"
    if args.vision_limit is None:
        args.vision_limit = 20
    if args.enable_vision is None:
        args.enable_vision = False
    if args.enable_faces is None:
        args.enable_faces = False
    if args.dry_run is None:
        args.dry_run = False
    if args.clear_cache is None:
        args.clear_cache = False

    return args

# ── Data model ─────────────────────────────────────────────────────────────────
@dataclass
class Photo:
    path: str
    filename: str = ""
    file_size: int = 0
    file_exists: bool = True
    # Inline metrics
    blur: float = 0.0
    brisque: float = 0.0
    composite_score: float = 0.0
    colorfulness: float = 0.0
    brightness: float = 0.0
    contrast: float = 0.0
    resolution: float = 0.0
    width: int = 0
    height: int = 0
    label: str = ""
    iso_speed: float = 0.0
    exposure_time_s: float = 0.0
    # Hashes
    md5: str = ""
    phash: str = ""
    # Groups
    dup_group: str = ""
    dup_rank: int = 0
    burst_group: str = ""
    burst_rank: int = 0
    is_burst_winner: bool = False
    # CLIP
    clip_tags: str = ""
    clip_confidence: float = 0.0
    # Event
    event_id: str = ""
    event_date: str = ""
    # Faces
    has_faces: bool = False
    face_count: int = 0
    person_ids: str = ""
    primary_face_bbox: str = ""
    # Vision LLM
    caption: str = ""
    vision_category: str = ""
    vision_quality: str = ""
    vision_memorability: int = 0
    vision_keep: str = ""
    vision_delete: str = ""
    vision_model: str = ""
    # Score & decision
    score: int = -1
    decision: str = ""
    reason: str = ""
    # Meta
    cache_key: str = ""
    error: str = ""


# ── Database ───────────────────────────────────────────────────────────────────
class DB:
    def __init__(self, path: str):
        self.conn = sqlite3.connect(path)
        self.conn.execute("PRAGMA journal_mode=WAL")
        self.conn.execute("PRAGMA synchronous=NORMAL")
        cols = []
        col_defs = {}
        for f in Photo.__dataclass_fields__.values():
            t = "BOOLEAN" if f.type in (bool, "bool") else \
                "INTEGER"  if f.type in (int, "int")   else \
                "REAL"     if f.type in (float, "float") else "TEXT"
            cols.append(f"{f.name} {t}")
            col_defs[f.name] = t
        cols[0] = "path TEXT PRIMARY KEY"
        self.conn.execute(f"CREATE TABLE IF NOT EXISTS photos ({', '.join(cols)})")
        existing = {r[1] for r in self.conn.execute("PRAGMA table_info(photos)").fetchall()}
        for col_name, col_type in col_defs.items():
            if col_name not in existing:
                self.conn.execute(f"ALTER TABLE photos ADD COLUMN {col_name} {col_type}")
        self.conn.commit()

    def save(self, p: Photo):
        d = asdict(p)
        cols = list(d.keys())
        sql = (f"INSERT INTO photos ({','.join(cols)}) VALUES ({','.join(['?']*len(cols))})"
               f" ON CONFLICT(path) DO UPDATE SET "
               f"{', '.join(f'{c}=excluded.{c}' for c in cols if c != 'path')}")
        self.conn.execute(sql, list(d.values()))
        self.conn.commit()

    def save_many(self, photos: List[Photo], commit_every: int = 100):
        if not photos:
            return
        for i, p in enumerate(photos, start=1):
            d = asdict(p)
            cols = list(d.keys())
            sql = (f"INSERT INTO photos ({','.join(cols)}) VALUES ({','.join(['?']*len(cols))})"
                   f" ON CONFLICT(path) DO UPDATE SET "
                   f"{', '.join(f'{c}=excluded.{c}' for c in cols if c != 'path')}")
            self.conn.execute(sql, list(d.values()))
            if i % commit_every == 0:
                self.conn.commit()
        self.conn.commit()

    def load(self, path: str) -> Optional[Photo]:
        cur = self.conn.execute("SELECT * FROM photos WHERE path=?", (path,))
        row = cur.fetchone()
        if row:
            cols = [d[0] for d in cur.description]
            try:
                return Photo(**dict(zip(cols, row)))
            except TypeError:
                return None
        return None

    def close(self):
        self.conn.close()


# ── Logging ────────────────────────────────────────────────────────────────────
def setup_log(output_dir: str) -> logging.Logger:
    log_path = Path(output_dir) / "engine.log"
    fmt = "%(asctime)s.%(msecs)03d %(levelname)s [%(name)s] %(message)s"
    datefmt = "%Y-%m-%d %H:%M:%S%z"
    logging.basicConfig(level=logging.INFO, format=fmt, handlers=[
        logging.FileHandler(log_path, encoding="utf-8"),
    ])
    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.INFO)
    console.setFormatter(logging.Formatter(fmt=fmt, datefmt=datefmt))
    for handler in logging.getLogger().handlers:
        if isinstance(handler, logging.FileHandler):
            handler.setFormatter(logging.Formatter(fmt=fmt, datefmt=datefmt))
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass
    logging.getLogger().addHandler(console)
    return logging.getLogger("photo")

log = logging.getLogger("photo")


@contextmanager
def stage_timer(label: str):
    started_at = datetime.now().astimezone()
    wall_tick = perf_counter()
    cpu_tick = process_time()
    log.info(f"[START] {label} @ {started_at.isoformat(timespec='milliseconds')}")
    try:
        yield
    finally:
        finished_at = datetime.now().astimezone()
        elapsed_wall = perf_counter() - wall_tick
        elapsed_cpu = process_time() - cpu_tick
        log.info(
            f"[END]   {label} @ {finished_at.isoformat(timespec='milliseconds')} "
            f"(elapsed_wall={elapsed_wall:.2f}s, elapsed_cpu={elapsed_cpu:.2f}s)"
        )


# ── Image metrics ──────────────────────────────────────────────────────────────
def compute_blur(path: Path) -> float:
    """Laplacian variance — higher = sharper."""
    try:
        if HAS_CV2:
            img = cv2.imread(str(path))
            if img is None:
                return -1.0
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            return float(cv2.Laplacian(gray, cv2.CV_64F).var())
        with Image.open(path) as img:
            px = list(img.convert("L").resize((256, 256)).getdata())
            m = sum(px) / len(px)
            return float(sum((p - m) ** 2 for p in px) / len(px))
    except Exception:
        return -1.0


def compute_brisque_approx(path: Path) -> float:
    """
    BRISQUE-like no-reference quality estimate using local normalized
    luminance coefficients (MSCN). Lower = better quality.
    Approximates OpenCV's BRISQUE without contrib module.
    Range: roughly 0-100, where <15 is excellent, >50 is poor.
    """
    try:
        if HAS_CV2:
            import numpy as np
            img = cv2.imread(str(path), cv2.IMREAD_GRAYSCALE)
            if img is None:
                return 50.0
            img = cv2.resize(img, (256, 256)).astype(np.float64)
            mu = cv2.GaussianBlur(img, (7, 7), 7 / 6)
            mu_sq = mu * mu
            sigma = cv2.GaussianBlur(img * img, (7, 7), 7 / 6)
            sigma = np.sqrt(np.abs(sigma - mu_sq)) + 1e-7
            mscn = (img - mu) / sigma
            mscn_flat = mscn.flatten()
            var = float(np.var(mscn_flat))
            kurt = float(np.mean(mscn_flat ** 4) / (var ** 2 + 1e-7)) - 3.0
            quality = 20.0 + abs(kurt) * 8.0 + abs(var - 1.0) * 15.0
            return max(0.0, min(100.0, quality))
        else:
            with Image.open(path) as img:
                gray = img.convert("L").resize((256, 256))
                if hasattr(gray, "get_flattened_data"):
                    px = list(gray.get_flattened_data())
                else:
                    px = list(gray.getdata())
            mean = sum(px) / len(px)
            var = sum((p - mean) ** 2 for p in px) / len(px)
            if var < 500:
                return min(100.0, 80.0 - var * 0.1)
            elif var > 4000:
                return min(100.0, 20.0 + (var - 4000) * 0.01)
            else:
                return max(5.0, 50.0 - (var - 500) * 0.012)
    except Exception:
        return 50.0


def compute_colorfulness(img: Image.Image) -> float:
    """Hasler & Susstrunk colorfulness metric."""
    try:
        rgb = img.convert("RGB").resize((256, 256))
        r_band = rgb.getchannel("R")
        g_band = rgb.getchannel("G")
        b_band = rgb.getchannel("B")
        if hasattr(r_band, "get_flattened_data"):
            r = list(r_band.get_flattened_data())
            g = list(g_band.get_flattened_data())
            b = list(b_band.get_flattened_data())
        else:
            r = list(r_band.getdata())
            g = list(g_band.getdata())
            b = list(b_band.getdata())
        n = len(r)
        rg = [r[i] - g[i] for i in range(n)]
        yb = [0.5 * (r[i] + g[i]) - b[i] for i in range(n)]
        rg_m = sum(rg) / n
        yb_m = sum(yb) / n
        rg_s = math.sqrt(sum((v - rg_m) ** 2 for v in rg) / n)
        yb_s = math.sqrt(sum((v - yb_m) ** 2 for v in yb) / n)
        return math.sqrt(rg_s ** 2 + yb_s ** 2) + 0.3 * math.sqrt(rg_m ** 2 + yb_m ** 2)
    except Exception:
        return 0.0


def compute_brightness_contrast(img: Image.Image) -> Tuple[float, float]:
    """Luminance mean (brightness) and std-dev (contrast)."""
    try:
        gray = img.convert("L").resize((256, 256))
        if hasattr(gray, "get_flattened_data"):
            px = list(gray.get_flattened_data())
        else:
            px = list(gray.getdata())
        mean = sum(px) / len(px)
        std = math.sqrt(sum((p - mean) ** 2 for p in px) / len(px))
        return mean, std
    except Exception:
        return 0.0, 0.0


def compute_composite(blur: float, brisque: float, resolution: float,
                       colorfulness: float, contrast: float) -> Tuple[float, str]:
    """
    Composite quality score (0-100k range for spread) and label.
    Combines all inline metrics into a single figure for relative ranking.
    """
    blur_norm = min(100.0, max(0.0, blur / 10.0)) if blur > 0 else 0.0
    brisque_norm = max(0.0, 100.0 - brisque)
    res_norm = min(100.0, resolution / 20000.0)
    color_norm = min(100.0, colorfulness)
    contrast_norm = min(100.0, contrast * 1.5)

    composite = (blur_norm * 0.35 +
                 brisque_norm * 0.25 +
                 res_norm * 0.15 +
                 color_norm * 0.15 +
                 contrast_norm * 0.10) * 1000.0

    if composite > 70000 and blur > 300:
        label = "GOOD"
    elif composite > 40000 and blur > 80:
        label = "AVERAGE"
    else:
        label = "POOR"

    return composite, label


# ── Helpers ────────────────────────────────────────────────────────────────────
def find_images(folder: str, limit: Optional[int]) -> List[Path]:
    imgs = set()
    for ext in SUPPORTED_EXTS:
        imgs.update(Path(folder).rglob(f"*{ext}"))
        imgs.update(Path(folder).rglob(f"*{ext.upper()}"))
    result = sorted(imgs)
    return result[:limit] if limit else result


def file_md5(path: Path) -> str:
    try:
        h = hashlib.md5()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""):
                h.update(chunk)
        return h.hexdigest()
    except Exception:
        return ""


def file_phash(path: Path) -> str:
    try:
        with Image.open(path) as img:
            return str(imagehash.phash(img.convert("RGB")))
    except Exception:
        return ""


def exif_ts(path: Path) -> Optional[datetime]:
    try:
        from PIL.ExifTags import TAGS
        with Image.open(path) as img:
            exif = img._getexif()
            if exif:
                for tag, val in exif.items():
                    if TAGS.get(tag) == "DateTimeOriginal":
                        return datetime.strptime(val, "%Y:%m:%d %H:%M:%S")
    except Exception:
        pass
    return None


def filename_ts(path: Path) -> Optional[datetime]:
    for pat in [r'(\d{4})(\d{2})(\d{2})_(\d{2})(\d{2})(\d{2})', r'(\d{4})(\d{2})(\d{2})']:
        m = re.search(pat, path.name)
        if m:
            try:
                g = list(map(int, m.groups()))
                return datetime(*g) if len(g) == 6 else datetime(g[0], g[1], g[2])
            except ValueError:
                pass
    return None


def get_ts(path: Path) -> datetime:
    return exif_ts(path) or filename_ts(path) or datetime.fromtimestamp(path.stat().st_mtime)


def to_b64(path: Path, max_size=(512, 512), face_bbox: str = "") -> str:
    try:
        resample = getattr(Image, 'LANCZOS', Image.Resampling.LANCZOS)
        with Image.open(path) as img:
            img = img.convert("RGB")
            w, h = img.size
            side = min(w, h)

            # LEFT tile: full-scene thumbnail for composition context
            left_tile = img.copy()
            left_tile.thumbnail(max_size, resample)
            left_canvas = Image.new("RGB", max_size, "black")
            left_canvas.paste(
                left_tile,
                ((max_size[0] - left_tile.width) // 2, (max_size[1] - left_tile.height) // 2)
            )

            # RIGHT tile: 1:1 native-detail center crop (300x300)
            center_crop = _safe_square_crop(img, w // 2, h // 2, side).resize((300, 300), resample)

            # Optional BOTTOM tile: primary face detail crop (300x300)
            face_tile = None
            if face_bbox:
                try:
                    t, r, b, l = [int(v) for v in face_bbox.split(",")]
                    fw = max(1, r - l)
                    fh = max(1, b - t)
                    fside = max(fw, fh) * 2
                    cx = l + fw // 2
                    cy = t + fh // 2
                    face_tile = _safe_square_crop(img, cx, cy, fside).resize((300, 300), resample)
                except Exception:
                    face_tile = None

            right_col_h = 600 if face_tile is not None else 300
            composite_h = max(max_size[1], right_col_h)
            composite = Image.new("RGB", (max_size[0] + 300, composite_h), "black")

            # Place left tile centered vertically.
            composite.paste(left_canvas, (0, (composite_h - max_size[1]) // 2))
            # Place right-side center detail tile.
            composite.paste(center_crop, (max_size[0], 0))
            # Place optional face tile on bottom-right.
            if face_tile is not None:
                composite.paste(face_tile, (max_size[0], 300))

            buf = BytesIO()
            composite.save(buf, format="JPEG", quality=88)
            return base64.b64encode(buf.getvalue()).decode()
    except Exception:
        return ""


def _safe_square_crop(img: Image.Image, cx: int, cy: int, size: int) -> Image.Image:
    w, h = img.size
    size = max(32, min(size, w, h))
    half = size // 2
    left = max(0, min(cx - half, w - size))
    top = max(0, min(cy - half, h - size))
    return img.crop((left, top, left + size, top + size))


def vision_multi_patch_b64(path: Path, face_bbox: str = "", panel_size: int = 672) -> str:
    """
    Build 3-panel composite:
      1) full scene (context)
      2) 1:1 center crop (texture)
      3) 1:1 primary-face crop if available, else center crop
    """
    try:
        resample = getattr(Image, 'LANCZOS', Image.Resampling.LANCZOS)
        with Image.open(path) as src:
            img = src.convert("RGB")
            w, h = img.size
            side = min(w, h)

            # Full scene panel
            full_panel = img.copy()
            full_panel.thumbnail((panel_size, panel_size), resample)
            canvas = Image.new("RGB", (panel_size, panel_size), "black")
            ox = (panel_size - full_panel.width) // 2
            oy = (panel_size - full_panel.height) // 2
            canvas.paste(full_panel, (ox, oy))
            full_panel = canvas

            # Center 1:1 panel
            center_panel = _safe_square_crop(img, w // 2, h // 2, side).resize((panel_size, panel_size), resample)

            # Primary-face 1:1 panel
            face_panel = center_panel
            if face_bbox:
                try:
                    t, r, b, l = [int(v) for v in face_bbox.split(",")]
                    fw = max(1, r - l)
                    fh = max(1, b - t)
                    fside = max(fw, fh) * 2
                    cx = l + fw // 2
                    cy = t + fh // 2
                    face_panel = _safe_square_crop(img, cx, cy, fside).resize((panel_size, panel_size), resample)
                except Exception:
                    pass

            composite = Image.new("RGB", (panel_size * 3, panel_size), "black")
            composite.paste(full_panel, (0, 0))
            composite.paste(center_panel, (panel_size, 0))
            composite.paste(face_panel, (panel_size * 2, 0))

            buf = BytesIO()
            composite.save(buf, format="JPEG", quality=88)
            return base64.b64encode(buf.getvalue()).decode()
    except Exception:
        return ""


def parse_json(raw: str) -> dict:
    raw = re.sub(r'^```(?:json)?\s*', '', raw.strip(), flags=re.MULTILINE)
    raw = re.sub(r'```\s*$', '', raw.strip(), flags=re.MULTILINE)
    raw = raw.strip()
    try:
        return json.loads(raw)
    except Exception:
        m = re.search(r'\{.*\}', raw, re.DOTALL)
        if m:
            try:
                return json.loads(m.group())
            except Exception:
                pass
    raise ValueError(f"Cannot parse JSON: {raw[:100]}")


def cache_key(folder: str, metrics_excel: Optional[str] = None) -> str:
    folder_abs = str(Path(folder).resolve())
    metrics_abs = str(Path(metrics_excel).resolve()) if metrics_excel else "inline_metrics"
    return hashlib.md5(f"{folder_abs}|{metrics_abs}".encode()).hexdigest()[:12]


def _to_float(val, default: float = 0.0) -> float:
    try:
        if val is None or val == "":
            return default
        return float(val)
    except Exception:
        return default


def _to_int(val, default: int = 0) -> int:
    try:
        if val is None or val == "":
            return default
        return int(val)
    except Exception:
        return default


def _norm_path(s: str) -> str:
    return str(s).strip().replace("\\", "/").lower()


def load_metrics_excel(path: str) -> Dict[str, dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    header_row = next(rows, None) or []
    headers = {str(h).strip().lower(): idx for idx, h in enumerate(header_row) if h is not None}

    by_path: Dict[str, dict] = {}
    by_filename: Dict[str, dict] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        path_val = row[headers["path"]] if "path" in headers else None
        file_val = row[headers["filename"]] if "filename" in headers else None
        if path_val is None and file_val is None:
            continue

        data = {
            "blur": _to_float(row[headers["blur"]], 0.0) if "blur" in headers else 0.0,
            "brisque": _to_float(row[headers["brisque"]], 0.0) if "brisque" in headers else 0.0,
            "composite_score": _to_float(row[headers["composite_score"]], 0.0) if "composite_score" in headers else 0.0,
            "label": str(row[headers["label"]]).strip().upper() if "label" in headers and row[headers["label"]] is not None else "",
            "resolution": _to_float(row[headers["resolution"]], 0.0) if "resolution" in headers else 0.0,
            "width": _to_int(row[headers["width"]], 0) if "width" in headers else 0,
            "height": _to_int(row[headers["height"]], 0) if "height" in headers else 0,
            "colorfulness": _to_float(row[headers["colorfulness"]], 0.0) if "colorfulness" in headers else 0.0,
            "brightness": _to_float(row[headers["brightness"]], 0.0) if "brightness" in headers else 0.0,
            "contrast": _to_float(row[headers["contrast"]], 0.0) if "contrast" in headers else 0.0,
        }

        if path_val:
            by_path[_norm_path(str(path_val))] = data
        if file_val:
            by_filename[str(file_val).strip().lower()] = data

    return {"by_path": by_path, "by_filename": by_filename}


# ── Stage 1: Load & compute all metrics inline ────────────────────────────────
def _match_metrics_row(img: Path, folder: str, metrics: Dict[str, dict]) -> Optional[dict]:
    by_path = metrics.get("by_path", {})
    by_filename = metrics.get("by_filename", {})

    abs_key = _norm_path(str(img))
    if abs_key in by_path:
        return by_path[abs_key]

    try:
        rel_key = _norm_path(str(img.relative_to(Path(folder))))
        if rel_key in by_path:
            return by_path[rel_key]
    except Exception:
        pass

    return by_filename.get(img.name.lower())




def _exif_num(v) -> float:
    try:
        if isinstance(v, tuple) and len(v) == 2 and v[1]:
            return float(v[0]) / float(v[1])
        return float(v)
    except Exception:
        return 0.0


def _extract_sensor_state(pil_img) -> tuple[float, float]:
    iso_speed = 0.0
    exposure_time_s = 0.0
    try:
        exif = pil_img.getexif()
        iso_speed = _exif_num(exif.get(34855, 0))  # ISOSpeedRatings
        exposure_time_s = _exif_num(exif.get(33434, 0))  # ExposureTime
    except Exception:
        pass
    return iso_speed, exposure_time_s

def compute_photo_metrics_worker(path_str: str, ck: str) -> dict:
    import numpy as np  # noqa: F401 - required by metric helpers in worker process
    img = Path(path_str)
    out = {
        "path": str(img),
        "filename": img.name,
        "file_size": 0,
        "cache_key": ck,
        "file_exists": True,
        "width": 0, "height": 0, "resolution": 0.0,
        "colorfulness": 0.0, "brightness": 0.0, "contrast": 0.0,
        "blur": 0.0, "brisque": 0.0, "composite_score": 0.0, "label": "",
        "iso_speed": 0.0, "exposure_time_s": 0.0,
        "error": "",
    }
    try:
        out["file_size"] = img.stat().st_size
    except Exception:
        out["file_exists"] = False
        out["error"] = "missing"
        return out
    try:
        with Image.open(img) as pil_img:
            out["width"], out["height"] = pil_img.size
            out["resolution"] = float(out["width"] * out["height"])
            out["iso_speed"], out["exposure_time_s"] = _extract_sensor_state(pil_img)
            rgb = pil_img.convert("RGB")
            out["colorfulness"] = round(compute_colorfulness(rgb), 2)
            brt, ctr = compute_brightness_contrast(rgb)
            out["brightness"] = round(brt, 2)
            out["contrast"] = round(ctr, 2)
    except Exception as e:
        out["error"] = f"PIL:{e}"

    out["blur"] = round(compute_blur(img), 2)
    out["brisque"] = round(compute_brisque_approx(img), 2)
    comp, label = compute_composite(
        out["blur"], out["brisque"], out["resolution"], out["colorfulness"], out["contrast"]
    )
    out["composite_score"] = round(comp, 2)
    out["label"] = label
    return out


def stage_load(images: List[Path], db: DB, ck: str,
               folder: str, metrics: Optional[Dict[str, dict]] = None) -> List[Photo]:
    log.info(f"S1 load+metrics: {len(images)} images")
    photos: List[Photo] = []
    pending: List[Path] = []
    pending_cached: List[Photo] = []
    for img in images:
        cached = db.load(str(img))
        if cached and cached.cache_key == ck and cached.md5:
            pending_cached.append(cached)
            continue
        pending.append(img)

    photos.extend(pending_cached)
    to_save_batch: List[Photo] = []
    max_workers = max(1, os.cpu_count() or 1)
    with ProcessPoolExecutor(max_workers=max_workers) as ex:
        futures = {ex.submit(compute_photo_metrics_worker, str(img), ck): img for img in pending}
        for fut in tqdm(as_completed(futures), total=len(futures), desc="S1 metrics"):
            img = futures[fut]
            try:
                result = fut.result()
            except Exception as e:
                p = Photo(path=str(img), filename=img.name, cache_key=ck, error=f"worker:{type(e).__name__}")
                photos.append(p)
                to_save_batch.append(p)
                if len(to_save_batch) >= 100:
                    db.save_many(to_save_batch, commit_every=100)
                    to_save_batch = []
                continue

            p = Photo(**result)
            if metrics:
                m = _match_metrics_row(img, folder, metrics)
                if m:
                    p.blur = round(m.get("blur", p.blur), 2)
                    p.brisque = round(m.get("brisque", p.brisque), 2)
                    p.composite_score = round(m.get("composite_score", p.composite_score), 2)
                    p.label = m.get("label") or p.label
                    if m.get("resolution", 0) > 0:
                        p.resolution = m.get("resolution", p.resolution)
                    if m.get("width", 0) > 0:
                        p.width = m.get("width", p.width)
                    if m.get("height", 0) > 0:
                        p.height = m.get("height", p.height)
                    if m.get("colorfulness", 0) > 0:
                        p.colorfulness = m.get("colorfulness", p.colorfulness)
                    if m.get("brightness", 0) > 0:
                        p.brightness = m.get("brightness", p.brightness)
                    if m.get("contrast", 0) > 0:
                        p.contrast = m.get("contrast", p.contrast)
            photos.append(p)
            to_save_batch.append(p)
            if len(to_save_batch) >= 100:
                db.save_many(to_save_batch, commit_every=100)
                to_save_batch = []

    if to_save_batch:
        db.save_many(to_save_batch, commit_every=100)

    log.info(f"             : {len(photos)} loaded with inline metrics")
    return photos


# ── Stage 2: Duplicates (MD5) ──────────────────────────────────────────────────
def stage_duplicates(photos: List[Photo], db: DB) -> List[Photo]:
    log.info("S2 duplicates: scanning")
    md5_groups = defaultdict(list)
    for p in tqdm(photos, desc="S2 MD5"):
        if not p.md5:
            p.md5 = file_md5(Path(p.path))
            db.save(p)
        if p.md5:
            md5_groups[p.md5].append(p)

    dup_count = 0
    for h, group in md5_groups.items():
        if len(group) > 1:
            group.sort(key=lambda x: (-x.composite_score, -x.blur, -x.file_size))
            for rank, p in enumerate(group):
                p.dup_group = f"DUP_{h[:8]}"
                p.dup_rank = rank + 1
                db.save(p)
            dup_count += 1
    log.info(f"             : {dup_count} duplicate groups")
    return photos


# ── Stage 3: Burst detection (pHash + timestamp) ──────────────────────────────
def stage_burst(photos: List[Photo], db: DB) -> List[Photo]:
    log.info("S3 burst     : scanning")
    valid = [p for p in photos if p.file_exists and not p.dup_group]
    for p in tqdm(valid, desc="S3 pHash"):
        if not p.phash:
            p.phash = file_phash(Path(p.path))
            db.save(p)

    # Pre-compute timestamps once (avoids millions of EXIF re-reads)
    log.info("             : computing timestamps")
    ts_map = {}
    for p in valid:
        ts_map[p.path] = get_ts(Path(p.path))

    # Sort by timestamp so we only compare nearby photos
    valid_sorted = sorted(valid, key=lambda p: ts_map[p.path])

    # Pre-parse hashes once
    hash_map = {}
    for p in valid_sorted:
        if p.phash:
            try:
                hash_map[p.path] = imagehash.hex_to_hash(p.phash)
            except Exception:
                pass

    groups, processed = [], set()
    for i, pa in enumerate(valid_sorted):
        if pa.path in processed or pa.path not in hash_map:
            continue
        ts_a = ts_map[pa.path]
        ha = hash_map[pa.path]
        group = [pa]
        processed.add(pa.path)
        # Only look forward in time-sorted list; stop when gap > 5 min
        for pb in valid_sorted[i + 1:]:
            if (ts_map[pb.path] - ts_a).total_seconds() > 300:
                break
            if pb.path in processed or pb.path not in hash_map:
                continue
            if ha - hash_map[pb.path] <= DUPLICATE_THRESH:
                group.append(pb)
                processed.add(pb.path)
        if len(group) > 1:
            groups.append(group)

    for gid, group in enumerate(groups):
        group.sort(key=lambda x: (-x.composite_score, -x.blur, -x.file_size))
        for rank, p in enumerate(group):
            p.burst_group = f"BURST_{gid:03d}"
            p.burst_rank = rank + 1
            p.is_burst_winner = (rank == 0)
            db.save(p)
    log.info(f"             : {len(groups)} burst groups")
    return photos


class GPUMonitor:
    def __init__(self, stage_name: str, enabled: bool = False, gpu_index: int = 0):
        self.stage_name = stage_name
        self.enabled = enabled
        self.gpu_index = gpu_index
        self.handle = None
        self.start = None
        self.samples = []

    def _snapshot(self):
        if not self.handle:
            return None
        try:
            util = nvmlDeviceGetUtilizationRates(self.handle)
            mem = nvmlDeviceGetMemoryInfo(self.handle)
            tx_kbs = nvmlDeviceGetPcieThroughput(self.handle, NVML_PCIE_UTIL_TX_BYTES)
            rx_kbs = nvmlDeviceGetPcieThroughput(self.handle, NVML_PCIE_UTIL_RX_BYTES)
            power_w = nvmlDeviceGetPowerUsage(self.handle) / 1000.0
            temp_c = nvmlDeviceGetTemperature(self.handle, NVML_TEMPERATURE_GPU)
            return {
                "gpu_util": float(util.gpu), "mem_util": float(util.memory),
                "vram_used": int(mem.used), "vram_free": int(mem.free), "vram_total": int(mem.total),
                "pcie_tx_kbs": float(tx_kbs), "pcie_rx_kbs": float(rx_kbs),
                "power_w": float(power_w), "temp_c": float(temp_c),
            }
        except Exception as e:
            if self.enabled:
                log.warning(f"[PERF] [{self.stage_name}] NVML snapshot failed: {type(e).__name__}")
            return None

    def sample(self):
        snap = self._snapshot()
        if snap:
            self.samples.append(snap)
        return snap

    def __enter__(self):
        if not (self.enabled and HAS_NVML):
            if self.enabled and not HAS_NVML:
                log.warning(f"[PERF] [{self.stage_name}] NVML unavailable; skipping GPU telemetry")
            return self
        try:
            nvmlInit()
            self.handle = nvmlDeviceGetHandleByIndex(self.gpu_index)
            self.start = self._snapshot()
            if self.start:
                self.samples.append(self.start)
        except Exception as e:
            self.handle = None
            log.warning(f"[PERF] [{self.stage_name}] NVML init failed: {type(e).__name__}")
        return self

    def __exit__(self, exc_type, exc, tb):
        if not self.enabled:
            return False
        end = self._snapshot() if self.handle else None
        if end:
            self.samples.append(end)
        if self.start and end:
            delta_vram_mb = (end["vram_used"] - self.start["vram_used"]) / (1024 * 1024)
            avg_sm = sum(x["gpu_util"] for x in self.samples) / max(1, len(self.samples))
            peak_tx = max(x["pcie_tx_kbs"] for x in self.samples) / 1024.0
            peak_rx = max(x["pcie_rx_kbs"] for x in self.samples) / 1024.0
            log.info(
                f"[PERF] [{self.stage_name}] Delta_VRAM:{delta_vram_mb:.1f}MB | "
                f"Avg_SM_Util:{avg_sm:.1f}% | Peak_PCIe_TX:{peak_tx:.1f}MB/s | Peak_PCIe_RX:{peak_rx:.1f}MB/s"
            )
        if self.handle:
            try:
                nvmlShutdown()
            except Exception:
                pass
        return False


def check_gpu_headroom(perf_log: bool = False, min_free_gb: float = 1.0) -> bool:
    if not HAS_NVML:
        return True
    try:
        nvmlInit()
        try:
            h = nvmlDeviceGetHandleByIndex(0)
            mem = nvmlDeviceGetMemoryInfo(h)
            free_gb = mem.free / (1024 ** 3)
            if free_gb < min_free_gb:
                log.warning(f"[PERF] [STAGE_7] Low GPU headroom: {free_gb:.2f}GB free; running cache cleanup")
                try:
                    import torch
                    if torch.cuda.is_available():
                        torch.cuda.empty_cache()
                except Exception:
                    pass
                gc.collect()
                return False
            return True
        finally:
            try:
                nvmlShutdown()
            except Exception:
                pass
    except Exception as e:
        if perf_log:
            log.warning(f"[PERF] [STAGE_7] GPU headroom check failed: {type(e).__name__}")
        return True


# ── Stage 4: CLIP tagging ─────────────────────────────────────────────────────
def stage_clip(photos: List[Photo], db: DB, perf_log: bool = False) -> List[Photo]:
    try:
        import clip
        import torch
    except ImportError:
        log.info("S4 CLIP      : skipped (not installed)")
        return photos

    try:
        device = "cuda" if torch.cuda.is_available() else "cpu"
        model, preprocess = clip.load("ViT-B/32", device=device)
        categories = ["indoor", "outdoor", "daytime", "nighttime", "people", "landscape",
                      "document", "food", "animal", "architecture", "screenshot", "event"]
        labels = ["indoor", "outdoor", "day", "night", "people", "landscape",
                  "document", "food", "animal", "architecture", "screenshot", "event"]
        tokens = clip.tokenize([f"a photo of {c}" for c in categories]).to(device)
        valid = [
            p for p in photos
            if p.file_exists and not p.clip_tags and p.resolution > 0 and not (p.error or "").startswith("worker:")
        ]
        log.info(f"S4 CLIP      : {len(valid)} images ({device})")
        with GPUMonitor("STAGE_4", enabled=perf_log) as gm:
            for p in tqdm(valid, desc="S4 CLIP"):
                try:
                    with Image.open(p.path) as img:
                        t = preprocess(img.convert("RGB")).unsqueeze(0).to(device)
                    with torch.no_grad():
                        lp = (100.0 * model.encode_image(t) @ model.encode_text(tokens).T).softmax(dim=-1)
                    probs = lp.cpu().numpy()[0]
                    top = [labels[i] for i in probs.argsort()[-3:][::-1] if probs[i] > 0.2]
                    p.clip_tags = ", ".join(top)
                    p.clip_confidence = float(max(probs))
                    if perf_log and device == "cuda":
                        try:
                            stats = torch.cuda.memory_stats()
                            reserved = stats.get("reserved_bytes.all.current", 0)
                            active = stats.get("active_bytes.all.current", 0)
                            frag_mb = (reserved - active) / (1024 * 1024)
                            if frag_mb > 500:
                                log.warning(f"[PERF] [STAGE_4] File: {Path(p.path).name} | VRAM_Fragmentation: {frag_mb:.1f}MB")
                        except Exception as e:
                            log.warning(f"[PERF] [STAGE_4] memory_stats failed: {type(e).__name__}")
                    gm.sample()
                    db.save(p)
                except Exception as e:
                    p.error = f"CLIP:{e}"
                    db.save(p)
    except Exception as e:
        log.error(f"S4 CLIP failed: {e}")
    return photos


# ── Stage 5: Event grouping ───────────────────────────────────────────────────
def stage_events(photos: List[Photo], db: DB) -> List[Photo]:
    valid = sorted(
        [(p, get_ts(Path(p.path))) for p in photos if p.file_exists],
        key=lambda x: x[1]
    )
    groups, current, last_ts = [], [], None
    for p, ts in valid:
        if last_ts is None or (ts - last_ts).total_seconds() / 3600 > EVENT_GAP_HOURS:
            if current:
                groups.append(current)
            current = []
        current.append((p, ts))
        last_ts = ts
    if current:
        groups.append(current)

    for gid, group in enumerate(groups):
        date_str = group[0][1].strftime("%Y-%m-%d")
        eid = f"{date_str}_{chr(65 + gid % 26)}"
        for p, ts in group:
            p.event_id = eid
            p.event_date = date_str
            db.save(p)
    log.info(f"S5 events    : {len(groups)} groups")
    return photos


# ── Stage 6: Face detection ───────────────────────────────────────────────────
def stage_faces(photos: List[Photo], db: DB) -> List[Photo]:
    try:
        import face_recognition
    except ImportError:
        log.info("S6 faces     : skipped (not installed)")
        return photos
    valid = [
        p for p in photos
        if p.file_exists and not p.has_faces and p.face_count == 0
        and p.resolution > 0 and not (p.error or "").startswith("worker:")
    ]
    log.info(f"S6 faces     : {len(valid)} images")
    all_encs = []
    for p in tqdm(valid, desc="S6 faces"):
        try:
            img = face_recognition.load_image_file(p.path)
            locs = face_recognition.face_locations(img)
            encs = face_recognition.face_encodings(img, locs)
            p.has_faces = len(locs) > 0
            p.face_count = len(locs)
            if locs:
                # Primary face = largest bounding box area
                primary = max(locs, key=lambda b: max(1, (b[1] - b[3]) * (b[2] - b[0])))
                p.primary_face_bbox = ",".join(str(v) for v in primary)  # top,right,bottom,left
            for e in encs:
                all_encs.append((e, p))
            db.save(p)
        except Exception as e:
            p.error = f"face:{e}"
            db.save(p)

    clusters = []
    for enc, p in all_encs:
        assigned = False
        for cid, cencs in clusters:
            if min(face_recognition.face_distance(cencs, enc)) < 0.55:
                cencs.append(enc)
                lbl = f"Person_{chr(65 + cid % 26)}"
                existing = p.person_ids.split(", ") if p.person_ids else []
                if lbl not in existing:
                    p.person_ids = ", ".join(existing + [lbl])
                assigned = True
                break
        if not assigned:
            cid = len(clusters)
            clusters.append((cid, [enc]))
            lbl = f"Person_{chr(65 + cid % 26)}"
            existing = p.person_ids.split(", ") if p.person_ids else []
            if lbl not in existing:
                p.person_ids = ", ".join(existing + [lbl])
        db.save(p)
    log.info(f"             : {len(clusters)} persons found")
    return photos


# ── Scoring engine ─────────────────────────────────────────────────────────────
def compute_scores(photos: List[Photo]) -> List[Photo]:
    """
    Score 0-100. Relative (folder percentiles) + absolute signals.

    Point budget:
      Blur           25 pts  (relative to folder p75)
      BRISQUE        15 pts  (absolute — lower is better)
      Composite      15 pts  (relative to folder p75)
      Label          10 pts  (GOOD/AVERAGE/POOR from composite)
      Burst winner   10 pts
      Event unique   10 pts
      Vision quality 10 pts  (excellent=+20, good=+12, avg=0, poor=-15)
      Memorability   15 pts  (1-5 scale from vision, semantic value)
    Total possible: ~100+ before clamping

    Hard overrides:
      Exact duplicate non-winner -> score=10, REMOVE
      Burst non-winner           -> score=20, REMOVE
    """
    # Folder percentiles for relative scoring
    blurs = sorted([p.blur for p in photos if p.blur > 0])
    composites = sorted([p.composite_score for p in photos if p.composite_score > 0])
    blur_p75 = blurs[int(len(blurs) * 0.75)] if blurs else 200.0
    comp_p75 = composites[int(len(composites) * 0.75)] if composites else 50000.0

    # Pre-compute event counts once
    event_counts = Counter(p.event_id for p in photos if p.event_id)

    LOW_LAPLACIAN_FLOOR = 30.0
    HIGH_ISO_CUTOFF = 1600.0
    FAST_SHUTTER_S = 1.0 / 1000.0
    LABEL_MAP   = {"GOOD": 10, "AVERAGE": 0, "POOR": -10}

    # Memorability: 1=mundane, 2=ordinary, 3=decent, 4=memorable, 5=exceptional
    # Maps to -5 / 0 / +5 / +10 / +15
    MEMO_MAP = {1: -5, 2: 0, 3: 5, 4: 10, 5: 15}

    for p in photos:
        s = 50.0

        # ── Hard overrides ────────────────────────────────────────────────
        if p.dup_group and p.dup_rank > 1:
            p.score = 10
            p.decision = "REMOVE"
            p.reason = f"Exact duplicate (copy #{p.dup_rank})"
            continue

        if p.burst_group and not p.is_burst_winner:
            p.score = 20
            p.decision = "REMOVE"
            p.reason = f"Burst duplicate (rank #{p.burst_rank} of group)"
            continue

        # ── Blur: relative to folder p75 (25 pts) ────────────────────────
        blur_good_floor = 300.0
        if 0 < p.exposure_time_s < FAST_SHUTTER_S:
            blur_good_floor = 220.0

        if p.blur > 0:
            ratio = p.blur / blur_p75
            if ratio >= 1.5:    s += 25
            elif ratio >= 1.0:  s += 15
            elif ratio >= 0.5:  s += 5
            elif ratio >= 0.2:  s -= 10
            else:               s -= 20

        # ── BRISQUE: absolute, lower is better (15 pts) ──────────────────
        if p.brisque > 0:
            noise_penalty_scale = 0.6 if p.iso_speed > HIGH_ISO_CUTOFF else 1.0
            if p.brisque < 15:    s += 15
            elif p.brisque < 30:  s += 8
            elif p.brisque < 50:  s += 2
            else:                 s -= 10 * noise_penalty_scale

        # ── Composite: relative to folder p75 (15 pts) ───────────────────
        if p.composite_score > 0:
            ratio = p.composite_score / comp_p75
            if ratio >= 1.5:    s += 15
            elif ratio >= 1.0:  s += 8
            elif ratio >= 0.5:  s += 2
            else:               s -= 5

        # ── Label (10 pts) ───────────────────────────────────────────────
        label_score = LABEL_MAP.get(p.label.upper() if p.label else "", 0)
        if p.label == "GOOD" and p.blur < blur_good_floor:
            label_score = 0
        s += label_score

        # ── Burst winner bonus (10 pts) ──────────────────────────────────
        if p.is_burst_winner:
            s += 10

        # ── Event uniqueness (10 pts) ────────────────────────────────────
        if p.event_id and event_counts[p.event_id] == 1:
            s += 10

        p.score = max(0, min(100, int(s)))

    # ── Assign decisions (Bayesian gate with VLM veto) ─────────────────────
    # Adaptive thresholds: only tighten when vision covered >50% of review band
    review_band = [p for p in photos if not p.decision and VISION_BAND_LOW <= p.score <= VISION_BAND_HIGH]
    vision_count = sum(1 for p in review_band if p.vision_model)
    vision_coverage = vision_count / max(len(review_band), 1)
    if vision_coverage > 0.5:
        keep_thresh, remove_thresh = 70, 30
    else:
        keep_thresh, remove_thresh = 65, 35

    
    high_value_categories = {"document", "people"}

    def deterministic_prior(photo: Photo) -> str:
        """Prior decision from deterministic signals only: blur/BRISQUE/resolution."""
        d = 50.0

        if photo.blur > 0:
            blur_ratio = photo.blur / blur_p75
            if blur_ratio >= 1.5:    d += 25
            elif blur_ratio >= 1.0:  d += 15
            elif blur_ratio >= 0.5:  d += 5
            elif blur_ratio >= 0.2:  d -= 10
            else:                    d -= 20

        if photo.brisque > 0:
            if photo.brisque < 15:    d += 15
            elif photo.brisque < 30:  d += 8
            elif photo.brisque < 50:  d += 2
            else:                     d -= 10

        if photo.resolution > 0:
            if photo.resolution >= 12_000_000:      d += 12
            elif photo.resolution >= 8_000_000:     d += 8
            elif photo.resolution >= 2_000_000:     d += 2
            elif photo.resolution < 1_000_000:      d -= 12

        d = max(0, min(100, int(d)))
        if d >= keep_thresh:
            return "KEEP"
        if d <= remove_thresh:
            return "REMOVE"
        return "REVIEW"


    for p in photos:
        if p.decision:
            continue

        prior_decision = deterministic_prior(p)

        if p.blur > 0 and p.blur < LOW_LAPLACIAN_FLOOR:
            prior_decision = "REMOVE"

        if prior_decision == "KEEP":
            p.decision = "KEEP"
            parts = []
            if p.is_burst_winner:                    parts.append("best in burst")
            if p.blur > blur_p75:                    parts.append("sharp")
            if p.brisque > 0 and p.brisque < 20:    parts.append("low noise")
            if p.label == "GOOD":                    parts.append("good quality")
            if p.vision_quality in ("excellent", "good"):
                parts.append(f"vision:{p.vision_quality}")
            if p.vision_memorability >= 4:
                parts.append(f"memorable({p.vision_memorability}/5)")
            p.reason = ", ".join(parts) if parts else "high quality score"

        elif prior_decision == "REMOVE":
            p.decision = "REMOVE"
            parts = []
            if p.blur > 0 and p.blur < BLUR_THRESH: parts.append("blurry")
            if p.brisque > 50:                       parts.append("high noise")
            if p.label == "POOR":                    parts.append("poor quality")
            if p.vision_quality == "poor":           parts.append("vision:poor")
            if p.vision_memorability == 1:           parts.append("mundane")
            p.reason = ", ".join(parts) if parts else "low quality score"

        else:
            p.decision = "REVIEW"
            p.reason = "ambiguous -- manual review suggested"

        # Bayesian veto gate: VLM quality can aggressively override borderline outcomes.
        category = (p.vision_category or "").strip().lower()
        quality = (p.vision_quality or "").strip().lower()
        if quality == "poor" and p.decision in ("REVIEW", "REMOVE"):
            p.decision = "REMOVE"
            p.reason = "VLM Veto: poor quality confirmed"
        elif quality == "excellent":
            p.decision = "KEEP"
            p.reason = "VLM Veto: excellent quality override"
        elif p.decision in ("REVIEW", "REMOVE") and quality == "good" and category in high_value_categories:
            p.decision = "KEEP"
            p.reason = "VLM Veto: High-value content despite low sharpness"

    scored = [p for p in photos if p.score >= 0]
    if scored:
        log.info(f"Scores       : min={min(p.score for p in scored)} "
                 f"max={max(p.score for p in scored)} "
                 f"avg={sum(p.score for p in scored) // len(scored)} "
                 f"| keep>={keep_thresh} remove<={remove_thresh}")
    return photos


# ── Stage 7: Vision LLM (selective, score band 35-65) ─────────────────────────
def stage_vision(photos: List[Photo], db: DB,
                 primary: str, fallback: str, limit: int, perf_log: bool = False) -> List[Photo]:
    candidates = [p for p in photos
                  if p.file_exists and not p.vision_model
                  and VISION_BAND_LOW <= p.score <= VISION_BAND_HIGH][:limit]

    if not candidates:
        log.info("S7 vision    : no candidates in review band")
        return photos

    log.info(f"S7 vision    : {len(candidates)} candidates (score {VISION_BAND_LOW}-{VISION_BAND_HIGH})")

    attempts = 0
    successes = 0
    fallback_successes = 0
    timeouts = 0
    json_fails = 0
    other_errors = 0
    api_wall_total = 0.0
    api_cpu_total = 0.0

    async def run_async() -> None:
        nonlocal attempts, successes, fallback_successes, timeouts, json_fails, other_errors, api_wall_total, api_cpu_total
        sem = asyncio.Semaphore(3)
        timeout_cfg = aiohttp.ClientTimeout(total=VISION_TIMEOUT)

        async with aiohttp.ClientSession(timeout=timeout_cfg) as session:
            async def process_one(p: Photo) -> None:
                nonlocal attempts, successes, fallback_successes, timeouts, json_fails, other_errors, api_wall_total, api_cpu_total
                check_gpu_headroom(perf_log=perf_log)
                encode_tick = perf_counter()
                b64 = to_b64(Path(p.path), max_size=(512, 512), face_bbox=p.primary_face_bbox)
                encode_secs = perf_counter() - encode_tick
                if not b64:
                    return

                async with sem:
                    for model in [primary, fallback]:
                        attempts += 1
                        try:
                            prompt = (
                                f"Blur={p.blur:.0f} brisque={p.brisque:.0f} composite={p.composite_score:.0f}. "
                                "Foveated image: LEFT scene context; RIGHT detail tiles for focus/noise. "
                                "Technical Audit Only. No conversational preamble. Output JSON and stop immediately. "
                                '{"caption":"one sentence","category":"people/landscape/food/document/animal/architecture/event/other",'
                                '"quality":"excellent/good/average/poor","memorability":1,"keep_reason":"...","delete_reason":"..."}'
                            )
                            api_wall_tick = perf_counter()
                            api_cpu_tick = process_time()
                            mon = GPUMonitor("STAGE_7", enabled=perf_log)
                            mon.__enter__()
                            pre_bus = mon._snapshot()
                            async with session.post(OLLAMA_URL, json={
                                "model": model, "prompt": prompt,
                                "images": [b64], "stream": False,
                                "options": {"temperature": 0.1, "num_predict": 220}
                            }) as resp:
                                resp.raise_for_status()
                                resp_json = await resp.json()
                            api_wall_total += perf_counter() - api_wall_tick
                            api_cpu_total += process_time() - api_cpu_tick
                            raw = resp_json.get("response", "").strip()
                            post_bus = mon._snapshot()
                            mon.__exit__(None, None, None)

                            data = parse_json(raw)
                            p.caption = data.get("caption", "")
                            p.vision_category = data.get("category", "")
                            p.vision_quality = data.get("quality", "")
                            p.vision_keep = data.get("keep_reason", "")
                            p.vision_delete = data.get("delete_reason", "")
                            memo = data.get("memorability", 0)
                            try:
                                p.vision_memorability = max(1, min(5, int(memo)))
                            except (ValueError, TypeError):
                                p.vision_memorability = 0

                            if perf_log:
                                b64_mb = len(b64.encode("utf-8")) / (1024 * 1024)
                                tx0 = (pre_bus or {}).get("pcie_tx_kbs", 0.0) / 1024.0
                                tx1 = (post_bus or {}).get("pcie_tx_kbs", 0.0) / 1024.0
                                bus_tx = max(tx0, tx1)
                                vram_gb = ((post_bus or {}).get("vram_used", 0.0)) / (1024 ** 3)
                                td = resp_json.get("total_duration", 0) / 1e6
                                ld = resp_json.get("load_duration", 0) / 1e6
                                ped = resp_json.get("prompt_eval_duration", 0) / 1e6
                                ed = resp_json.get("eval_duration", 0) / 1e6
                                log.info(
                                    f"[PERF] [STAGE_7] File: {Path(p.path).name} | Bus_TX: {bus_tx:.1f}MB/s | "
                                    f"Base64_Size: {b64_mb:.2f}MB | Encode_Ship: {encode_secs*1000:.1f}ms | "
                                    f"VRAM_Used: {vram_gb:.2f}GB | Ollama_Total: {td:.1f}ms | "
                                    f"Ollama_Load: {ld:.1f}ms | Ollama_Prompt: {ped:.1f}ms | Ollama_Eval: {ed:.1f}ms"
                                )
                            p.vision_model = model
                            db.save(p)
                            successes += 1
                            if model != primary:
                                fallback_successes += 1
                            return
                        except (ValueError, json.JSONDecodeError) as e:
                            log.warning(f"Vision JSON fail {Path(p.path).name} ({model}): {e}")
                            p.error = f"vision_json:{model}"
                            json_fails += 1
                        except asyncio.TimeoutError:
                            log.warning(f"Vision timeout {Path(p.path).name} ({model})")
                            p.error = f"vision_timeout:{model}"
                            timeouts += 1
                        except Exception as e:
                            log.warning(f"Vision error {Path(p.path).name} ({model}): {type(e).__name__}")
                            p.error = f"vision_err:{model}"
                            other_errors += 1

            tasks = [asyncio.create_task(process_one(p)) for p in candidates]
            for fut in tqdm(asyncio.as_completed(tasks), total=len(tasks), desc="S7 vision"):
                await fut

    asyncio.run(run_async())

    photos = compute_scores(photos)
    processed = sum(1 for p in photos if p.vision_model)
    elapsed_minutes = max(api_wall_total / 60.0, 1e-9)
    concurrent_throughput = successes / elapsed_minutes
    avg_wall = (api_wall_total / successes) if successes else 0.0
    avg_cpu = (api_cpu_total / successes) if successes else 0.0
    log.info(
        "             : "
        f"{processed} processed | attempts={attempts} success={successes} "
        f"fallback_success={fallback_successes} timeout={timeouts} "
        f"json_fail={json_fails} error={other_errors} "
        f"| vision_avg_wall={avg_wall:.2f}s vision_avg_cpu={avg_cpu:.2f}s "
        f"Concurrent_Throughput={concurrent_throughput:.2f} img/min"
    )
    return photos


# ── Output: Excel ──────────────────────────────────────────────────────────────
def write_excel(photos: List[Photo], path: str):
    COLS = [
        ("filename",        "Filename",   28, "left"),
        ("score",           "Score /100",  10, "center"),
        ("decision",        "Decision",   12, "center"),
        ("vision_category", "Category",   14, "center"),
        ("caption",         "Caption",    50, "left"),
        ("reason",          "Reason",     35, "left"),
    ]

    DECISION_STYLE = {
        "KEEP":   ("006100", "C6EFCE"),
        "REMOVE": ("9C0006", "FFC7CE"),
        "REVIEW": ("7D6608", "FFEB9C"),
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Photo Cleanup"

    thin = Border(
        bottom=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD")
    )

    for ci, (_, label, width, _) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=label)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Sort strictly by score ascending (worst first)
    sorted_photos = sorted(
        [p for p in photos if p.file_exists],
        key=lambda p: (p.score, p.filename.lower())
    )

    alt_fill = [
        PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid"),
    ]

    for ri, p in enumerate(sorted_photos, 2):
        bg = alt_fill[ri % 2]
        for ci, (fld, _, _, align) in enumerate(COLS, 1):
            val = getattr(p, fld, "") or ""
            if fld == "score":
                val = int(val) if val != "" else ""
            # CLIP tags as fallback category when vision hasn't run
            if fld == "vision_category" and not val and p.clip_tags:
                val = p.clip_tags
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = thin
            c.alignment = Alignment(
                horizontal=align, vertical="center",
                wrap_text=(fld in ("caption", "reason"))
            )

            if fld == "decision" and val in DECISION_STYLE:
                fg, bg_col = DECISION_STYLE[val]
                c.font = Font(name="Arial", bold=True, color=fg, size=10)
                c.fill = PatternFill(start_color=bg_col, end_color=bg_col, fill_type="solid")
            elif fld == "score":
                score = int(val) if val != "" else 50
                if score >= 70:     col = "C6EFCE"
                elif score >= 36:   col = "FFEB9C"
                else:               col = "FFC7CE"
                c.font = Font(name="Arial", size=10, bold=True)
                c.fill = PatternFill(start_color=col, end_color=col, fill_type="solid")
            else:
                c.font = Font(name="Arial", size=10)
                c.fill = bg

        ws.row_dimensions[ri].height = 32 if p.caption else 18

    # Summary footer
    ws.append([])
    sr = ws.max_row + 1
    keep = sum(1 for p in sorted_photos if p.decision == "KEEP")
    remove = sum(1 for p in sorted_photos if p.decision == "REMOVE")
    review = sum(1 for p in sorted_photos if p.decision == "REVIEW")
    for i, (lbl, val) in enumerate([
        ("Total", len(sorted_photos)),
        ("KEEP", keep), ("REMOVE", remove), ("REVIEW", review)
    ]):
        ws.cell(row=sr + i, column=1, value=lbl).font = Font(name="Arial", bold=True)
        ws.cell(row=sr + i, column=2, value=val).font = Font(name="Arial", bold=True)

    wb.save(path)
    log.info(f"Excel        : KEEP={keep} REMOVE={remove} REVIEW={review}")


# ── Output: CSV ────────────────────────────────────────────────────────────────
def write_csv(photos: List[Photo], path: str):
    order = {"REMOVE": 0, "REVIEW": 1, "KEEP": 2}
    rows = sorted(
        [p for p in photos if p.file_exists],
        key=lambda p: (order.get(p.decision, 3), p.score)
    )
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["filename", "score", "decision", "category", "caption", "reason", "path"])
        for p in rows:
            w.writerow([p.filename, p.score, p.decision,
                        p.vision_category or p.clip_tags,
                        p.caption, p.reason, p.path])
    log.info(f"CSV          : {len(rows)} rows")


def write_metrics_excel(photos: List[Photo], path: str, folder: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Metrics"
    headers = [
        "path", "filename", "blur", "brisque", "composite_score", "label",
        "resolution", "width", "height", "colorfulness", "brightness", "contrast"
    ]
    ws.append(headers)
    root = Path(folder)
    for p in photos:
        full = Path(p.path)
        try:
            rel = str(full.relative_to(root)).replace("/", "\\")
        except Exception:
            rel = p.path
        ws.append([
            rel, p.filename, p.blur, p.brisque, p.composite_score, p.label,
            p.resolution, p.width, p.height, p.colorfulness, p.brightness, p.contrast
        ])
    wb.save(path)
    log.info(f"Metrics XLSX : {len(photos)} rows -> {path}")


# ── Dry run ────────────────────────────────────────────────────────────────────
def print_dry_run(photos: List[Photo]):
    buckets = {"0-20": 0, "21-35": 0, "36-50": 0, "51-65": 0, "66-80": 0, "81-100": 0}
    for p in photos:
        s = p.score
        if s <= 20:     buckets["0-20"] += 1
        elif s <= 35:   buckets["21-35"] += 1
        elif s <= 50:   buckets["36-50"] += 1
        elif s <= 65:   buckets["51-65"] += 1
        elif s <= 80:   buckets["66-80"] += 1
        else:           buckets["81-100"] += 1

    keep = sum(1 for p in photos if p.decision == "KEEP")
    remove = sum(1 for p in photos if p.decision == "REMOVE")
    review = sum(1 for p in photos if p.decision == "REVIEW")

    print("\n-- DRY RUN: Score Distribution -----------------------------")
    for band, count in buckets.items():
        bar = "#" * (count * 40 // max(len(photos), 1))
        print(f"  {band:8s} |{bar:<40s}| {count}")
    print(f"\n  KEEP: {keep}  REMOVE: {remove}  REVIEW: {review}  TOTAL: {len(photos)}")
    print("  (no files written in dry-run mode)\n")


def _bytes_to_gb(num_bytes: int) -> float:
    return round(num_bytes / (1024 ** 3), 1)


def _total_memory_bytes() -> Optional[int]:
    if hasattr(os, "sysconf"):
        try:
            pages = os.sysconf("SC_PHYS_PAGES")
            page_size = os.sysconf("SC_PAGE_SIZE")
            return int(pages * page_size)
        except Exception:
            return None
    return None


def _detect_gpu() -> Tuple[bool, str]:
    nvidia_smi = shutil.which("nvidia-smi")
    if not nvidia_smi:
        return False, "No NVIDIA GPU detected (nvidia-smi unavailable)"
    try:
        out = subprocess.check_output(
            [nvidia_smi, "--query-gpu=name,memory.total", "--format=csv,noheader"],
            stderr=subprocess.STDOUT,
            text=True,
            timeout=3,
        ).strip()
        if not out:
            return False, "nvidia-smi returned no GPU rows"
        first = out.splitlines()[0]
        return True, f"NVIDIA GPU: {first}"
    except Exception as exc:
        return False, f"GPU probe failed: {exc}"


def detect_runtime_profile(args) -> Dict[str, str]:
    mem_bytes = _total_memory_bytes()
    has_gpu, gpu_detail = _detect_gpu()
    clip_available = importlib.util.find_spec("clip") is not None and importlib.util.find_spec("torch") is not None
    face_available = importlib.util.find_spec("face_recognition") is not None
    profile = {
        "host": platform.node() or "unknown",
        "os": f"{platform.system()} {platform.release()} ({platform.machine()})",
        "python": sys.version.split()[0],
        "cpu_cores": str(os.cpu_count() or "unknown"),
        "ram": f"{_bytes_to_gb(mem_bytes)} GB" if mem_bytes else "unknown",
        "gpu": gpu_detail,
        "opencv": "available" if HAS_CV2 else "missing (fallback blur estimator)",
        "clip_lib": "available" if clip_available else "missing",
        "face_lib": "available" if face_available else "missing",
        "vision_requested": "yes" if args.enable_vision else "no",
        "vision_model": args.vision_model,
        "vision_fallback": args.vision_fallback,
    }

    if args.enable_vision:
        try:
            resp = requests.get("http://localhost:11434/api/tags", timeout=2)
            names = [m.get("name", "") for m in resp.json().get("models", [])]
            names_str = ", ".join(names[:10]) if names else "(none)"
            profile["ollama"] = "reachable"
            profile["ollama_models"] = names_str
            profile["vision_model_ready"] = "yes" if any(n.startswith(args.vision_model) for n in names) else "no"
            profile["vision_fallback_ready"] = "yes" if any(n.startswith(args.vision_fallback) for n in names) else "no"
        except Exception as exc:
            profile["ollama"] = f"unreachable ({exc})"
            profile["ollama_models"] = "unknown"
            profile["vision_model_ready"] = "unknown"
            profile["vision_fallback_ready"] = "unknown"
    return profile


def print_runtime_profile(profile: Dict[str, str]):
    print("\n-- DRY RUN: Hardware + Model Capability Analysis -----------")
    print(f"  Host             : {profile['host']}")
    print(f"  OS               : {profile['os']}")
    print(f"  Python           : {profile['python']}")
    print(f"  CPU cores        : {profile['cpu_cores']}")
    print(f"  RAM              : {profile['ram']}")
    print(f"  GPU              : {profile['gpu']}")
    print(f"  OpenCV           : {profile['opencv']}")
    print(f"  CLIP library     : {profile['clip_lib']}")
    print(f"  Face library     : {profile['face_lib']}")
    print(f"  Vision requested : {profile['vision_requested']}")
    print(f"  Vision model     : {profile['vision_model']}")
    if "ollama" in profile:
        print(f"  Ollama           : {profile['ollama']}")
        print(f"  Models           : {profile['ollama_models']}")
        print(f"  Model ready      : {profile['vision_model_ready']}")
        print(f"  Fallback ready   : {profile['vision_fallback_ready']}")
    print("")


def print_runtime_estimates(photos: List[Photo], args, profile: Dict[str, str]):
    total = len(photos)
    ambiguous = sum(1 for p in photos if VISION_BAND_LOW <= p.score <= VISION_BAND_HIGH)
    vision_calls = min(ambiguous, args.vision_limit) if args.enable_vision else 0
    cpu_cores = os.cpu_count() or 4
    has_gpu = profile.get("gpu", "").startswith("NVIDIA GPU")

    per_image = 0.06 if cpu_cores <= 4 else 0.045 if cpu_cores <= 8 else 0.03
    if not HAS_CV2:
        per_image += 0.015
    clip_cost = 0.02 if HAS_CLIP else 0.0
    vision_cost = 2.5 if has_gpu else 6.0
    output_cost = 0.25 if args.dry_run else 1.5

    est_secs = total * (per_image + clip_cost) + (vision_calls * vision_cost) + output_cost
    low = max(1, int(est_secs * 0.7))
    high = max(low + 1, int(est_secs * 1.35))

    print("-- DRY RUN: Runtime Estimate -------------------------------")
    print(f"  Photos scanned          : {total}")
    print(f"  Ambiguous (score 35-65) : {ambiguous}")
    print(f"  Vision calls planned    : {vision_calls} (limit={args.vision_limit})")
    print(f"  Estimated runtime       : {low}s to {high}s")
    print("  Notes                   : Range includes I/O variance, cache hits, and model warm-up.\n")


def print_teaser_results(photos: List[Photo]):
    ranked = sorted([p for p in photos if p.file_exists], key=lambda p: p.score)
    if not ranked:
        print("-- DRY RUN: Teaser Results --------------------------------")
        print("  No valid images found to preview.\n")
        return

    print("-- DRY RUN: Teaser Results --------------------------------")
    for p in ranked[:5]:
        print(f"  REMOVE-candidate teaser : score={p.score:3d}  file={p.filename}  reason={p.reason or 'low-quality signal'}")
    print("")


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description=f"Photo Cleanup Engine v{VERSION}")
    parser.add_argument("--folder", required=True,
                        help="Path to photo folder")
    parser.add_argument("--config", default="config.json",
                        help="Path to JSON config (default: config.json)")
    parser.add_argument("--output-dir", default=None,
                        help="Output directory (default: from config or output)")
    parser.add_argument("--limit", type=int, default=None,
                        help="Process only N images (default: from config or all)")
    parser.add_argument("--enable-vision", action="store_true", default=None,
                        help="Use Ollama vision LLM on ambiguous photos")
    parser.add_argument("--enable-faces", action="store_true", default=None,
                        help="Enable face detection and clustering")
    parser.add_argument("--vision-model", default=None,
                        help="Primary Ollama vision model (default: from config or llava)")
    parser.add_argument("--vision-fallback", default=None,
                        help="Fallback vision model (default: from config or llama3.2-vision)")
    parser.add_argument("--vision-limit", type=int, default=None,
                        help="Max images to send to vision LLM (default: from config or 20)")
    parser.add_argument("--metrics-excel", default=None,
                        help="Path to metrics Excel file (optional)")
    parser.add_argument("--generate-metrics", action="store_true",
                        help="Generate metrics Excel from folder scan and exit")
    parser.add_argument("--dry-run", action="store_true", default=None,
                        help="Preview score distribution, write no files")
    parser.add_argument("--clear-cache", action="store_true", default=None,
                        help="Wipe cached results before running")
    parser.add_argument("--perf-log", action="store_true",
                        help="Enable high-fidelity GPU performance instrumentation")
    args = parser.parse_args()
    config = _read_config(args.config)
    args = _apply_config_defaults(args, config)

    if args.enable_vision:
        num_parallel = OLLAMA_NUM_PARALLEL
        if num_parallel < 4:
            raise RuntimeError("OLLAMA_NUM_PARALLEL must be set to at least 4 for concurrent vision inference.")

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    global log
    log = setup_log(str(out_dir))

    if args.clear_cache:
        db_path = out_dir / "cache.db"
        if db_path.exists():
            db_path.unlink()
        log.info("Cache        : cleared")

    start = datetime.now()
    start_cpu = process_time()
    ck = cache_key(args.folder, args.metrics_excel)

    log.info(f"Photo Cleanup Engine v{VERSION}")
    log.info(f"Folder       : {args.folder}")
    log.info(f"Config       : limit={args.limit or 'all'} vision={args.enable_vision} faces={args.enable_faces}")
    if args.dry_run:
        profile = detect_runtime_profile(args)
        print_runtime_profile(profile)
    else:
        profile = {}

    metrics = None
    if args.metrics_excel:
        mpath = Path(args.metrics_excel)
        if not mpath.exists():
            raise FileNotFoundError(f"--metrics-excel not found: {args.metrics_excel}")
        metrics = load_metrics_excel(str(mpath))
        log.info(f"Metrics      : loaded from {args.metrics_excel}")

    db = DB(str(out_dir / "cache.db"))
    stale_error_count = db.conn.execute(
        "SELECT COUNT(*) FROM photos WHERE (error LIKE 'worker:%' OR resolution=0) AND cache_key<>?",
        (ck,)
    ).fetchone()[0]
    if stale_error_count and not args.clear_cache:
        raise RuntimeError(
            f"Found {stale_error_count} stale/corrupted cached rows from prior runs. "
            f"Re-run with --clear-cache to avoid cascading failures in CLIP/face stages."
        )

    try:
        with stage_timer("Scan images"):
            images = find_images(args.folder, args.limit)
        log.info(f"Found        : {len(images)} images")

        with stage_timer("S1 load+metrics"):
            photos = stage_load(images, db, ck, args.folder, metrics)

        if args.generate_metrics:
            metrics_name = f"{Path(args.folder).name}_metrics.xlsx"
            with stage_timer("Write metrics workbook"):
                write_metrics_excel(photos, str(out_dir / metrics_name), args.folder)
            log.info("Done         : metrics generated")
            return

        with stage_timer("S2 duplicates"):
            photos = stage_duplicates(photos, db)
        with stage_timer("S3 burst detection"):
            photos = stage_burst(photos, db)
        with stage_timer("S4 CLIP tagging"):
            photos = stage_clip(photos, db, perf_log=args.perf_log)
        with stage_timer("S5 event grouping"):
            photos = stage_events(photos, db)
        if args.enable_faces:
            with stage_timer("S6 face detection"):
                photos = stage_faces(photos, db)

        # Initial scoring before vision
        with stage_timer("Compute scores"):
            photos = compute_scores(photos)

        if args.enable_vision:
            with stage_timer("S7 vision analysis"):
                photos = stage_vision(photos, db,
                                      args.vision_model, args.vision_fallback,
                                      args.vision_limit, perf_log=args.perf_log)

        if args.dry_run:
            with stage_timer("Dry-run summary"):
                print_dry_run(photos)
                print_runtime_estimates(photos, args, profile)
                print_teaser_results(photos)
        else:
            with stage_timer("Write Excel"):
                write_excel(photos, str(out_dir / "photo_cleanup.xlsx"))
            with stage_timer("Write CSV"):
                write_csv(photos, str(out_dir / "photo_cleanup.csv"))

        wall_secs = (datetime.now() - start).total_seconds()
        cpu_secs = process_time() - start_cpu
        log.info(f"Done         : {len(photos)} photos in wall={wall_secs:.0f}s cpu={cpu_secs:.0f}s")

    finally:
        db.close()


if __name__ == "__main__":
    main()
